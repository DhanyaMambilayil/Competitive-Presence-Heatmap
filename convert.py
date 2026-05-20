#!/usr/bin/env python3
"""
convert.py
Reads branches.xlsx and writes branches.json for the heatmap.
Run: python convert.py
"""

import json, sys
try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not found.  Run:  pip install openpyxl")
    sys.exit(1)

REQUIRED_COLS = ["Company","Branch Name","Area","Governorate","Type",
                 "Latitude","Longitude","Address","Phone","Hours","Active"]
VALID_COMPANIES = {"BEC","Al Muzaini","Al Mulla","Al Ansari"}
VALID_TYPES     = {"Branch","Kiosk"}

def convert(xlsx_path="branches.xlsx", json_path="branches.json"):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Branches"]

    # Read header row
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    missing = [h for h in REQUIRED_COLS if h not in col]
    if missing:
        print(f"ERROR: Missing columns in Excel: {missing}")
        sys.exit(1)

    rows = []
    errors = []
    skipped = 0

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue  # skip blank rows

        def get(field):
            v = row[col[field]]
            return str(v).strip() if v is not None else ""

        active = get("Active").lower()
        if active != "yes":
            skipped += 1
            continue

        company = get("Company")
        name    = get("Branch Name")
        area    = get("Area")
        gov     = get("Governorate")
        btype   = get("Type")
        addr    = get("Address")
        phone   = get("Phone")
        hours   = get("Hours")

        # Validate
        row_errors = []
        if company not in VALID_COMPANIES:
            row_errors.append(f"Unknown company '{company}' (valid: {VALID_COMPANIES})")
        if btype not in VALID_TYPES:
            row_errors.append(f"Unknown type '{btype}' (valid: {VALID_TYPES})")
        if not name:
            row_errors.append("Branch Name is empty")

        try:
            lat = float(row[col["Latitude"]])
            lng = float(row[col["Longitude"]])
        except (TypeError, ValueError):
            row_errors.append(f"Invalid lat/lng: {row[col['Latitude']]}, {row[col['Longitude']]}")
            lat = lng = 0.0

        if not (28.5 <= lat <= 30.0):
            row_errors.append(f"Latitude {lat} looks wrong for Kuwait (expected 28.5–30.0)")
        if not (46.5 <= lng <= 49.0):
            row_errors.append(f"Longitude {lng} looks wrong for Kuwait (expected 46.5–49.0)")

        if row_errors:
            errors.append(f"Row {r_idx} ({name}): " + "; ".join(row_errors))
            continue

        rows.append({
            "n":   name,
            "co":  company.lower().replace(" ","_").replace("_",""),   # bec / almuzaini / almulla / alansari
            "company": company,
            "area": area,
            "gov":  gov,
            "type": btype.lower(),   # branch / kiosk
            "lat":  round(lat, 6),
            "lng":  round(lng, 6),
            "address": addr,
            "phone":   phone,
            "hours":   hours,
        })

    # Normalise company key
    co_map = {
        "BEC":        "bec",
        "Al Muzaini": "muzaini",
        "Al Mulla":   "mulla",
        "Al Ansari":  "ansari",
    }
    for r in rows:
        r["co"] = co_map.get(r["company"], r["co"])

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)

    print(f"✅  Converted {len(rows)} active locations → {json_path}")
    if skipped:
        print(f"⏭️   Skipped {skipped} inactive rows (Active ≠ Yes)")
    if errors:
        print(f"\n⚠️  {len(errors)} validation error(s):")
        for e in errors:
            print(f"   {e}")
    return len(rows), errors

if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser(description="Convert branches.xlsx → branches.json")
    p.add_argument("--xlsx", default="branches.xlsx", help="Input Excel file")
    p.add_argument("--json", default="branches.json", help="Output JSON file")
    args = p.parse_args()
    convert(args.xlsx, args.json)
